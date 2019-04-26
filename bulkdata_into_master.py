'''
Programme to move bulk data into the master spreadsheet

Input documents:
1) spreadsheet containing the information to place into the spreadsheet
2) master spreadsheet to be altered

Output document:
1) altered master spreadsheet (although there is the option to create a new document first to test that data is being
moved over correctly)

Instructions:
1) file path to spreadsheet containing information
2) file path to master to be altered
3) file path to master (to over write master)
'''

from openpyxl import load_workbook
from collections import OrderedDict
from openpyxl.utils import column_index_from_string
import datetime
from openpyxl.styles import Font
from bcompiler.utils import project_data_from_master


def put_data_in_dictionary(worksheet):
    d_dict = {}
    for row in worksheet.iter_rows(min_row=2):
        tasks_name = ""
        o = OrderedDict()
        for cell in row:
            if cell.column == 'A':
                tasks_name = cell.value
                d_dict[tasks_name] = o
            else:
                val = worksheet.cell(row=1, column=column_index_from_string(cell.column)).value
                if type(cell.value) == datetime:
                    d_value = datetime(cell.value.year, cell.value.month, cell.value.day)
                    d_dict[tasks_name][val] = d_value
                else:
                    d_dict[tasks_name][val] = cell.value
    try:
        del d_dict[None]
    except KeyError:
        pass

    return d_dict


def put_into_master(dictionary, worksheet):
    red_text = Font(color="00fc2525")
    for col_num in range(2, worksheet.max_column + 1):
        project_name = ws.cell(row=1, column=col_num).value
        print(project_name)
        if project_name in dictionary:
                for row_num in range(2, worksheet.max_row + 1):
                    for item in dictionary[project_name].keys():
                        if ws.cell(row=row_num, column=1).value == item:
                            if ws.cell(row=row_num, column=col_num).value == dictionary[project_name][item]:
                                pass
                            else:
                                ws.cell(row=row_num, column=col_num).value = dictionary[project_name][item]
                                ws.cell(row=row_num, column=col_num).font = red_text

    return wb


'''1) Specify the file path to spreadsheet containing information'''
data = project_data_from_master('C:\\Users\\Standalone\\Will\\Q4 DCA Ratings and VfM Data for GMPP Projects.xlsx')

'''2) Specify the file path to master'''
wb = load_workbook(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_4_2018_wip.xlsx')
ws = wb.active

amended_master = put_into_master(data, ws)

'''
3) File path to document being save here, whicb is the same as quarter master data path above.
NOTE: this effectively overwrites the master document so make sure you have saved the master before running this 
programme. If you wanted to create a document that doesn't overwrite the master and check changes first, you can type 
a different filename (such as 'test'). However, you will need to save all changes into/overwrite the master at some 
point - as it is the sole source of persistent final data.
'''
amended_master.save('C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_4_2018_wip.xlsx')
