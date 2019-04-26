'''
Programme changes order of master data into the another usable format e.g. internal master format into gmpp master
format and vice versa.

input documents
1) data_to_be_changed master = master data
2) data_map = datamap setting out the structure for the new format.

output document
1) excel master file with  data re-ordered as per the datamap.

see supplementary instructions on how to work with dms
'''

from openpyxl import load_workbook, Workbook
from bcompiler.utils import project_data_from_master

def create_master(data, dm):
    ws = dm.active

    for i, name in enumerate(data):
        print(name)
        ws.cell(row=1, column=2+i).value = name  # place project names in file

        for row_num in range(2, ws.max_row+1):
            key = ws.cell(row=row_num, column=1).value

            if key in data[name].keys():
                ws.cell(row=row_num, column=2+i).value = data[name][key]

    return dm

'''1) file path to master containing data for which the order is to be changed'''

master_order_to_be_altered = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\core data\\"
                                      "master_4_2018.xlsx")

'''2) file path to dm setting out structure for the new format'''
datamap = load_workbook("C:\\Users\\Standalone\\Will\\masters folder\\dms\\internal_dm_excel_master_for_merging.xlsx")

output = create_master(master_order_to_be_altered, datamap)

'''3) file path and name for new master that is being created'''
output.save("C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_4_2018_internal_format.xlsx")