'''programme for determining whether master keys have changed over time

developing this to help with the change over in reporting template

in development'''


#from openpyxl import load_workbook
from openpyxl import Workbook
#from collections import OrderedDict
#from openpyxl.utils import column_index_from_string
#import datetime
#from openpyxl.styles import Font
from bcompiler.utils import project_data_from_master

def get_keys(dictionary):
    wb = Workbook()
    ws = wb.active

    get = list(dictionary['Crossrail Programme'].keys())

    for x in range(0, len(get)):
        ws.cell(row=x + 2, column=1, value=get[x])

    return wb

data = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_4_2018.xlsx')

run = get_keys(data)

run.save('C:\\Users\\Standalone\\Will\\get_keys.xlsx')